#!/usr/bin/env python3
"""Generate sample.xlsx with 3 sheets for XLSX parser tests."""

from pathlib import Path
from openpyxl import Workbook

wb = Workbook()

# Sheet 1: Summary (3 rows, 4 cols)
ws1 = wb.active
ws1.title = "Summary"
ws1.append(["Category", "Q1", "Q2", "Q3"])
ws1.append(["Revenue", "1.2M", "1.5M", "1.8M"])
ws1.append(["Costs", "0.8M", "0.9M", "1.0M"])

# Sheet 2: Budget (4 rows, 3 cols)
ws2 = wb.create_sheet("Budget")
ws2.append(["Dept", "Alloc", "Spent"])
ws2.append(["Engineering", "500K", "320K"])
ws2.append(["Design", "200K", "150K"])
ws2.append(["Marketing", "150K", "140K"])

# Sheet 3: Empty (headers only, no data rows)
ws3 = wb.create_sheet("Empty")
ws3.append(["Column A", "Column B"])

out = Path(__file__).parent / "sample.xlsx"
wb.save(out)
print(f"Wrote {out} ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")
